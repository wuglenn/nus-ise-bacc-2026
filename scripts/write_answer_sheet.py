"""
Write output CSVs into a copy of the BACC 2026 Participant Answer Sheet.

Usage:
  python scripts/write_answer_sheet.py
"""

from __future__ import annotations

import csv
import shutil
from pathlib import Path

import openpyxl
from verify_constraints import QUARTERS, flow_cell, tooling_cell

ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "input" / "5) BACC 2026 Participant Answer Sheet.xlsx"
OUTPUT_DIR = ROOT / "output"
SOLUTION_DIR = ROOT / "solution"
SOLUTION_XLSX = SOLUTION_DIR / "5) BACC 2026 Participant Answer Sheet.xlsx"


def _quarter_index(quarter: str) -> int:
    return QUARTERS.index(quarter)


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="") as f:
        reader = csv.DictReader(f)
        return [row for row in reader if any(row.values())]


def _write_tooling(sheet, csv_path: Path) -> None:
    rows = _read_csv(csv_path)
    for row in rows:
        q_idx = _quarter_index(row["quarter"])
        ws = row["ws"]
        for fab in (1, 2, 3):
            value = int(float(row[f"fab{fab}"]))
            cell = tooling_cell(q_idx, fab, ws)
            sheet[cell].value = value


def _write_flow(sheet, csv_path: Path) -> None:
    rows = _read_csv(csv_path)
    if not rows:
        return
    for row in rows:
        q_idx = _quarter_index(row["quarter"])
        node = int(row["node"])
        step = int(row["step"])
        fab = int(row["fab"])
        loading = float(row["loading"])
        cell = flow_cell(q_idx, node, step, fab)
        sheet[cell].value = loading


def _write_q2(sheet, csv_path: Path, node: int) -> None:
    rows = _read_csv(csv_path)
    if not rows:
        return
    base_row = {1: 56, 2: 61, 3: 66}[node]
    for row in rows:
        q_idx = _quarter_index(row["quarter"])
        col = 4 + q_idx  # Column D = Q1'26
        sheet.cell(base_row + 0, col).value = float(row["expected_loading"])
        sheet.cell(base_row + 1, col).value = float(row["combined_variance"])
        sheet.cell(base_row + 2, col).value = float(row["prob_undercap_scen1"])


def main() -> None:
    SOLUTION_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(INPUT_XLSX, SOLUTION_XLSX)

    wb = openpyxl.load_workbook(SOLUTION_XLSX)

    q1a = wb["Q1a"]
    q1b = wb["Q1b"]
    q2 = wb["Input Data + Q2"]

    _write_tooling(q1a, OUTPUT_DIR / "01_q1a_tooling.csv")
    _write_flow(q1a, OUTPUT_DIR / "02_q1a_flow.csv")

    _write_tooling(q1b, OUTPUT_DIR / "03_q1b_tooling.csv")
    _write_flow(q1b, OUTPUT_DIR / "04_q1b_flow.csv")

    _write_q2(q2, OUTPUT_DIR / "05_q2_node1.csv", node=1)
    _write_q2(q2, OUTPUT_DIR / "06_q2_node2.csv", node=2)
    _write_q2(q2, OUTPUT_DIR / "07_q2_node3.csv", node=3)

    wb.save(SOLUTION_XLSX)


if __name__ == "__main__":
    main()
